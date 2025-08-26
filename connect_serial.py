import serial
from serial.tools import list_ports
import threading
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

# PORT = "/dev/cu.usbmodem214401"
BAUDRATE = 115200

ports = list_ports.comports()
for p in ports:
    print(p.device)
    print(p.product)
    print(p.description)
    print(p.serial_number)
    if p.serial_number == "75425D4EE3BBBAEA":
        print("Found Arduino Nano 33 BLE")
        PORT = p.device

ser = serial.Serial(PORT, BAUDRATE)

# File settings
file_name = "arduino_data.xlsx"
save_interval = 1  # Save every 1 records
data = []  # Buffer for data to save periodically

# Initialize the Excel file if it doesn't exist
if not os.path.exists(file_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Timestamp", "targetGapstate0","Gapstate0", "LEDstate0","targetGapstate1", "Gapstate1", "LEDstate1","targetGapstate2", "Gapstate2", "LEDstate2","targetGapstate3", "Gapstate3", "LEDstate3","targetGapstate4", "Gapstate4", "LEDstate4"])  # Add headers
    workbook.save(file_name)

def send_data(text):
    ser.write(str(text).encode("utf-8"))


def listen_to_serial(ser, data):
    """
    This function listens to the serial port and prints any received lines.
    """
    while True:
        line = ser.readline().decode('utf-8').strip()  # Read and decode line
        if line:
            print(f"Received: {line}")
            # timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S:%f')
             # Split the received line into separate columns
            values = line.split(',')  # Adjust delimiter if necessary
            # row = [timestamp] + values  # Combine timestamp with received data
            # data.append(row)
            data.append(values)

            # Save to Excel at regular intervals
            if len(data) >= save_interval:
                workbook = load_workbook(file_name)
                sheet = workbook.active
                
                for record in data:
                    sheet.append(record)  # Append rows directly to the sheet
                
                workbook.save(file_name)
                workbook.close()
                print(f"Saved {len(data)} records to {file_name}")
                
                # Clear saved data from memory
                data = []


def listen_to_serial_in_thread(ser,data):
    """
    This function listens to the serial port in a separate thread.
    """
    listener_thread = threading.Thread(target=listen_to_serial, args=(ser,data,))
    listener_thread.daemon = True
    listener_thread.start()

    return listener_thread


def main():
    # Start the listener thread
    listener_thread = listen_to_serial_in_thread(ser,data)
    print("Listening to serial port: " + PORT)

    # Keep the main thread alive to keep listening
    try:
        while True:
            user_input = input("Enter message to send (or 'exit' to quit): ")
            if user_input.lower() == "exit":
                break
            send_data(user_input)
    except KeyboardInterrupt:
        print("Exiting...")
    finally:
        print("Waiting for listener thread to finish...")
        listener_thread.join()  # Wait for the listener thread to clean up
        ser.close()
        print("Serial connection closed.")


if __name__ == "__main__":
    main()
